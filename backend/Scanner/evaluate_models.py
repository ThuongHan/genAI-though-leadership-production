"""
evaluate_models.py — per-dimension agreement between LLM scores and human scores.

Compares one or more model_scores_*.json files (from score_dataset.py) against the
human annotations in the v7 Annotate sheet, on each of the 4 rubric dimensions
(relevance, trustworthiness, specificity, audience_fit) plus the weighted grade.

METRICS (per dimension, computed only on articles BOTH the human and the model
scored — an inner join on annotation_id):

  Spearman rho   scipy.stats.spearmanr  — rank correlation: do they ORDER the same?
  QWK            sklearn.metrics.cohen_kappa_score(weights="quadratic")
                   — chance-corrected agreement on the 0-10 scale, big gaps cost more
  Kendall tau-b  scipy.stats.kendalltau — tie-robust rank correlation (robustness check)
  MAE            sklearn.metrics.mean_absolute_error — avg points apart, in score units
  Bias (signed)  mean(model - human) — does the model inflate (+) or deflate (-)?
  % within 1     share where |model - human| <= 1
  % exact        share where model == human

USAGE
-----
    python3 -m backend.Scanner.evaluate_models
    python3 -m backend.Scanner.evaluate_models backend/Scanner/data/eval/model_scores_claude_claude-sonnet-4-6_zero_shot.json
    python3 -m backend.Scanner.evaluate_models --annotations "backend/Scanner/data/annotation/dataset_kickstartAI_v7_final (2).xlsx"
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import sys
from pathlib import Path

# Add backend/Scanner/ to sys.path so `from main_scanner.xxx` imports resolve
# when running as `python3 -m backend.Scanner.evaluate_models` from pipeline/.
_SCANNER = Path(__file__).resolve().parent
sys.path.insert(0, str(_SCANNER))

from scipy.stats import kendalltau, spearmanr
from sklearn.metrics import cohen_kappa_score, mean_absolute_error

# Same dimensions + weights as the scorer, so the weighted grade is consistent.
from main_scanner.weighted_filter import RUBRIC_DIMS, RUBRIC_MAX, WEIGHTS

DEFAULT_ANNOTATIONS = str(_SCANNER / "data" / "annotation" / "dataset_kickstartAI_final scored.xlsx")
DEFAULT_GLOB        = str(_SCANNER / "data" / "eval" / "model_scores_*.json")
OUTPUT_CSV          = str(_SCANNER / "data" / "eval" / "eval_summary.csv")
ANNOTATE_SHEET = "Annotate"

# Map each dimension to the substring identifying its column header in the xlsx
# (headers look like "Relevance\n(0-10)", "Audience Fit\n(0-10)").
_HEADER_HINT = {
    "relevance": "relevance",
    "trustworthiness": "trustworthiness",
    "specificity": "specificity",
    "audience_fit": "audience",
}


# ---------------------------------------------------------------------------
# Metrics — thin wrappers over scipy / sklearn (+ trivial inline ones)
# ---------------------------------------------------------------------------

def _safe(fn, *a):
    """Run a stats fn, returning nan on the degenerate cases (e.g. no variance)."""
    try:
        v = fn(*a)
        return float(v) if v == v else float("nan")   # pass nan through
    except Exception:
        return float("nan")


def spearman_rho(human, model) -> float:
    return _safe(lambda h, m: spearmanr(h, m).correlation, human, model)


def kendall_tau(human, model) -> float:
    return _safe(lambda h, m: kendalltau(h, m).correlation, human, model)


def qwk(human, model) -> float:
    """Quadratic-weighted Cohen's kappa over the fixed 0..RUBRIC_MAX label set
    (fixing labels keeps it stable even if a score level is unused in a batch)."""
    labels = list(range(RUBRIC_MAX + 1))
    return _safe(
        lambda h, m: cohen_kappa_score([int(round(x)) for x in h],
                                       [int(round(x)) for x in m],
                                       labels=labels, weights="quadratic"),
        human, model,
    )


def bias(human, model) -> float:
    """Mean signed error: positive => model scores HIGHER than the human."""
    return sum(m - h for h, m in zip(human, model)) / len(human)


def pct_within(human, model, tol: float) -> float:
    return 100.0 * sum(1 for h, m in zip(human, model) if abs(h - m) <= tol) / len(human)


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def load_human_scores(xlsx_path: str) -> dict[int, dict[str, float]]:
    """Read the Annotate sheet -> {annotation_id: {dim: score}} for filled rows."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed (pip install openpyxl)")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[ANNOTATE_SHEET] if ANNOTATE_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(h or "").lower() for h in rows[0]]

    def col_for(hint: str) -> int | None:
        for i, h in enumerate(header):
            if hint in h:
                return i
        return None

    id_col = col_for("annotation_id")
    dim_cols = {dim: col_for(hint) for dim, hint in _HEADER_HINT.items()}
    if id_col is None or any(c is None for c in dim_cols.values()):
        print(f"ERROR: could not locate expected columns in {xlsx_path}")
        print(f"  headers seen: {rows[0]}")
        sys.exit(1)

    out: dict[int, dict[str, float]] = {}
    for row in rows[1:]:
        aid = row[id_col]
        if aid is None:
            continue
        scores: dict[str, float] = {}
        for dim, c in dim_cols.items():
            v = row[c]
            if isinstance(v, (int, float)):
                scores[dim] = float(v)
        if scores:                              # keep rows with at least one score
            out[int(aid)] = scores
    return out


def load_model_scores(json_path: str) -> tuple[str, dict[int, dict[str, float]]]:
    """Read a model_scores_*.json -> (label, {annotation_id: {dim: score}})."""
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    label = f"{data.get('provider', '?')}/{data.get('model', '?')}/{data.get('strategy', '?')}"
    out: dict[int, dict[str, float]] = {}
    for row in data.get("scored", []):
        aid = row.get("annotation_id")
        if aid is None or row.get("score_error"):
            continue
        out[int(aid)] = {d: float(row["scores"][d]) for d in RUBRIC_DIMS if d in row["scores"]}
    return label, out


def weighted_of(scores: dict[str, float]) -> float | None:
    """Mission-weighted average if all 4 dims present, else None."""
    if not all(d in scores for d in RUBRIC_DIMS):
        return None
    total_w = sum(WEIGHTS.values())
    return sum(scores[d] * WEIGHTS[d] for d in RUBRIC_DIMS) / total_w


# ---------------------------------------------------------------------------
# Evaluation
# ---------------------------------------------------------------------------

def paired(human, model, dim: str) -> tuple[list[float], list[float]]:
    """Inner-join human & model on annotation_id for one dimension (aligned lists)."""
    h_out, m_out = [], []
    for aid, hs in human.items():
        ms = model.get(aid)
        if ms is None or dim not in hs or dim not in ms:
            continue
        h_out.append(hs[dim])
        m_out.append(ms[dim])
    return h_out, m_out


def paired_weighted(human, model) -> tuple[list[float], list[float]]:
    h_out, m_out = [], []
    for aid, hs in human.items():
        ms = model.get(aid)
        if ms is None:
            continue
        hw, mw = weighted_of(hs), weighted_of(ms)
        if hw is None or mw is None:
            continue
        h_out.append(hw)
        m_out.append(mw)
    return h_out, m_out


def evaluate_one(label: str, human, model) -> list[dict]:
    """Return one metrics row per dimension (+ weighted) for a single model."""
    rows = []
    for dim in list(RUBRIC_DIMS) + ["weighted"]:
        h, m = paired_weighted(human, model) if dim == "weighted" else paired(human, model, dim)
        n = len(h)
        if n == 0:
            rows.append({"model": label, "dimension": dim, "n": 0})
            continue
        row = {
            "model": label,
            "dimension": dim,
            "n": n,
            "spearman": spearman_rho(h, m),
            "kendall": kendall_tau(h, m),
            "mae": float(mean_absolute_error(h, m)),
            "bias": bias(h, m),
            "pct_within1": pct_within(h, m, 1),
            "pct_exact": pct_within(h, m, 0),
            # QWK needs integer categories — meaningful for the per-dim 0-10 scores,
            # not the continuous weighted grade.
            "qwk": qwk(h, m) if dim != "weighted" else None,
        }
        rows.append(row)
    return rows


def _fmt(v) -> str:
    if v is None:
        return "   -"
    if isinstance(v, float):
        return "  nan" if v != v else f"{v:6.3f}"
    return f"{v:>4}"


def print_report(label: str, rows: list[dict]) -> None:
    print("\n" + "=" * 86)
    print(f"MODEL: {label}")
    print("=" * 86)
    print(f"{'dimension':<16}{'n':>4}{'spearman':>10}{'kendall':>9}{'qwk':>8}"
          f"{'mae':>8}{'bias':>8}{'within1':>9}{'exact':>8}")
    print("-" * 86)
    for r in rows:
        if r["n"] == 0:
            print(f"{r['dimension']:<16}{0:>4}   (no annotated rows yet)")
            continue
        print(f"{r['dimension']:<16}{r['n']:>4}{_fmt(r['spearman']):>10}{_fmt(r['kendall']):>9}"
              f"{_fmt(r.get('qwk')):>8}{_fmt(r['mae']):>8}{_fmt(r['bias']):>8}"
              f"{_fmt(r['pct_within1']):>9}{_fmt(r['pct_exact']):>8}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("files", nargs="*", help="model_scores_*.json (default: all in data/eval/)")
    parser.add_argument("--annotations", default=DEFAULT_ANNOTATIONS,
                        help="Human-annotated xlsx (default: the v7 final sheet)")
    parser.add_argument("--output", default=OUTPUT_CSV, help=f"CSV out (default: {OUTPUT_CSV})")
    args = parser.parse_args()

    if not Path(args.annotations).exists():
        print(f"ERROR: annotations file not found: {args.annotations}")
        sys.exit(1)

    human = load_human_scores(args.annotations)
    print(f"Human annotations: {len(human)} row(s) with at least one score "
          f"(from {Path(args.annotations).name})")
    if not human:
        print("\nNo human scores filled in yet — nothing to compare. "
              "Run again once the company returns the annotated sheet.")
        return

    files = args.files or sorted(glob.glob(DEFAULT_GLOB))
    if not files:
        print(f"\nNo model score files found ({DEFAULT_GLOB}). Run score_dataset.py first.")
        return

    all_rows: list[dict] = []
    for path in files:
        label, model = load_model_scores(path)
        rows = evaluate_one(label, human, model)
        print_report(label, rows)
        all_rows.extend(rows)

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    cols = ["model", "dimension", "n", "spearman", "kendall", "qwk", "mae", "bias",
            "pct_within1", "pct_exact"]
    with open(args.output, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in all_rows:
            w.writerow({c: r.get(c, "") for c in cols})
    print(f"\nWrote per-dimension metrics -> {args.output}")


if __name__ == "__main__":
    main()

# python3 -m backend.Scanner.evaluate_models 